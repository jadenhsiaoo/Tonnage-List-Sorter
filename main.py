from datetime import date, timedelta
import openpyxl
import module
import excel_module

today = date.today()
dayofweek1 = today.weekday()
dayofweek = int(dayofweek1)


#Monday
if dayofweek == 0:
    
    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb['Tonnage List']

    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=3), today - timedelta(days=4)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)


#Tuesday
if dayofweek == 1:
    
    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb['Tonnage List']

    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=5), today - timedelta(days=1), today - timedelta(days = 4)] + future_days 

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)

#Wednesday
if dayofweek == 2:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active
    
    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1)] + future_days
    rows = module.filter_rows(ws, days)
    
    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)


#Thursday
if dayofweek == 3:
    
    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today -timedelta(days=3)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)


#Friday
if dayofweek == 4:
    
    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today - timedelta(days=3), today - timedelta(days=4)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)

#Saturday
if dayofweek == 5:
    
    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today - timedelta(days=3), today - timedelta(days=4), today - timedelta(days=5)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)

#Sunday
if dayofweek == 6:
    
    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today - timedelta(days=3), today - timedelta(days=4), today - timedelta(days=5), today - timedelta(days= 6)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            module.txtNCHINA(rows)
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            module.txtSCHINA(rows)
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            module.txtSEA(rows)
            excel_module.excelSEA(rows)
        elif region == 'PG':
            module.txtPG(rows)
            excel_module.excelPG(rows)
        elif region == 'MED':
            module.txtMED(rows)
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            module.txtECSA(rows)
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            module.txtATL(rows)
            excel_module.excelATL(rows)