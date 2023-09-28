from datetime import date
import openpyxl


today = date.today()
dayofweek1 = today.weekday()
dayofweek = int(dayofweek1)

wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
ws = wb.active

today1 = str(date.today())

# Function to filter rows by region
def filter_rows(ws, days):
  rows_by_region = {
    'NCHINA': [], 
    'SCHINA': [],
    'SEA': [],
    'PG': [],
    'MED': [], 
    'ECSA': [],
    'BALTIC': []
  }

  for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
    
    try:
        total = row[0].value
        if not total or total.date() not in days:
            continue

    except Exception:
        continue

    if row[8].value in ['N.CHN', 'KOR.JPN', 'FEAST', 'WORLDWIDE', 'NOPAC', 'PNW']:
        rows_by_region['NCHINA'].append(row)
  
    elif row[8].value == 'S.CHN':
        rows_by_region['SCHINA'].append(row)
    
    elif row[8].value == 'SEA':
        rows_by_region['SEA'].append(row)

    elif row[8].value in ['PG-IND', 'PMO', 'SAF']:
        rows_by_region['PG'].append(row)

    elif row[8].value in ['MED', 'BSEA']:
        rows_by_region['MED'].append(row)
    
    elif row[8].value == 'ECSA':
        rows_by_region['ECSA'].append(row)
    
    elif row[8].value in ['BALTIC', 'NCSA', 'UK-CONTI', 'USEC', 'WAF', 'USG']:
        rows_by_region['BALTIC'].append(row)

  return rows_by_region



def format_update(value):
    if value is None:
        return 'N/A'
    value1 = str(value)
    month = int(value1[5:7])
    day = int(value1[8:10])
    if month == 1:
        return f"{day}-Jan"
    if month == 2:
        return f"{day}-Feb"
    if month == 3:
        return f"{day}-Mar"
    if month == 4:
        return f"{day}-Apr"
    if month == 5:
        return f"{day}-May"
    if month == 6:
        return f"{day}-Jun"
    if month == 7:
        return f"{day}-Jul"
    if month == 8:
        return f"{day}-Aug"
    if month == 9:
        return f"{day}-Sep"
    if month == 10:
        return f"{day}-Oct"
    if month == 11: 
        return f"{day}-Nov"
    if month == 12:
        return f"{day}-Dec"


def format_DWT(value):
    if value is None:
        return f"N/A"
    value1 = str(value)
    if len(value1) == 5:
        front = value1[:2]
        back = value1[2:]
        return f"{front},{back}"
    if len(value1) == 6:
        front = value1[:3]
        back = value1[3:]
        return f"{front},{back}"


def format_layday(value):
    if value is None:
        return f'N/A'
    value1 = str(value)
    year = value1[2:4]
    month = int(value1[5:7])
    day = value1[8:10]
    if month == 1:
        return f"{day}-Jan-{year}"
    if month == 2:
        return f"{day}-Feb-{year}"
    if month == 3:
        return f"{day}-Mar-{year}"
    if month == 4:
        return f"{day}-Apr-{year}"
    if month == 5:
        return f"{day}-May-{year}"
    if month == 6:
        return f"{day}-Jun-{year}"
    if month == 7:
        return f"{day}-Jul-{year}"
    if month == 8:
        return f"{day}-Aug-{year}"
    if month == 9:
        return f"{day}-Sep-{year}"
    if month == 10:
        return f"{day}-Oc-{year}t"
    if month == 11:
        return f"{day}-Nov-{year}"
    if month == 12:
        return f"{day}-Dec-{year}"

labels = ['UPDATE', 'VESSEL', 'DWT', 'BUILT', 'DRAFT', 'DOP', 'LAYDAY', 'OWNER/BROKER']  


def txtNCHINA(sortedlist):
    with open('Tonnage_List.txt', 'w') as file:
        file.write(today1)
        file.write('\n')
        file.write('\n')
        file.write('NCN-KOR-JPN')
        file.write('\n')
        file.write('\n')  
        file.write("{:<10}".format('UPDATE'))
        file.write("{:<23}".format('VESSEL'))
        file.write("{:<12}".format('DWT'))
        file.write("{:<10}".format('BUILT'))
        file.write("{:<10}".format('DRAFT'))
        file.write("{:<17}".format('DOP'))
        file.write("{:<12}".format('LAYDAY'))
        file.write("{:<20}".format('OWNER/BROKER'))
        file.write('\n')
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            # Write the cell values to the file with formatting
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')



def txtSCHINA(sortedlist):
    with open('Tonnage_List.txt', 'a') as file:
        file.write('\n')  
        file.write('S.CHINA')
        file.write('\n')
        file.write('\n') 
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            # Write the cell values to the file with formatting
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')


def txtSEA(sortedlist):
    with open('Tonnage_List.txt', 'a') as file:  
        file.write('\n')
        file.write('SEA')
        file.write('\n')
        file.write('\n')  # move this line out of the for loop to avoid new line for each label
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')


def txtPG(sortedlist):
    with open('Tonnage_List.txt', 'a') as file:
        file.write('\n')  
        file.write('PG-IND-SAF')
        file.write('\n')
        file.write('\n')  # move this line out of the for loop to avoid new line for each label
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            # Write the cell values to the file with formatting
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')



def txtMED(sortedlist):
    with open('Tonnage_List.txt', 'a') as file:
        file.write('\n') 
        file.write('MED')
        file.write('\n')
        file.write('\n')  # move this line out of the for loop to avoid new line for each label
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            # Write the cell values to the file with formatting
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')


def txtECSA(sortedlist):
    with open('Tonnage_List.txt', 'a') as file: 
        file.write('\n')
        file.write('ECSA')
        file.write('\n')
        file.write('\n')  # move this line out of the for loop to avoid new line for each label
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            # Write the cell values to the file with formatting
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')



def txtATL(sortedlist):
    with open('Tonnage_List.txt', 'a') as file:
        file.write('\n') 
        file.write('ATLANTIC')
        file.write('\n')
        file.write('\n')  # move this line out of the for loop to avoid new line for each label
        for row in sortedlist:
            # Get the first 8 cells (columns) of the row
            first_eight_columns = row[:8]
            # Write the cell values to the file with formatting
            for i, c in enumerate(first_eight_columns):
                if i == 0 and c.value is not None: #Update 
                    formatted_value = format_update(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 0 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 1 and c.value is not None: #Vessel Name
                    formatted_value = str(c.value)
                    file.write("{:<23}".format(formatted_value))
                elif i == 1 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<23}".format(formatted_value))
                elif i == 2 and c.value is not None: #DWT
                    formatted_value = str(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 2 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 3 and c.value is not None: #Built 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 3 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is not None: #Draft 
                    formatted_value = str(c.value)
                    file.write("{:<10}".format(formatted_value))
                elif i == 4 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<10}".format(formatted_value))
                elif i == 5 and c.value is not None: #DOP
                    formatted_value = str(c.value)
                    file.write("{:<17}".format(formatted_value))
                elif i == 5 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<17}".format(formatted_value))
                elif i == 6 and c.value is not None: #LAYDAY
                    formatted_value = format_layday(c.value)
                    file.write("{:<12}".format(formatted_value))
                elif i == 6 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<12}".format(formatted_value))
                elif i == 7 and c.value is not None: #Owner/Broker
                    formatted_value = str(c.value)
                    file.write("{:<20}".format(formatted_value))
                elif i == 7 and c.value is None:
                    formatted_value = "N/A"
                    file.write("{:<20}".format(formatted_value))
            file.write('\n')


