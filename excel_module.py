from datetime import date
import openpyxl
import module

today = date.today()
dayofweek1 = today.weekday()
dayofweek = int(dayofweek1)

wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
ws = wb.active

today1 = str(date.today())



def excelNCHINA(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active

    total_rows = ws_output.max_row
    ws_output.delete_rows(1, total_rows)
    ws_output.append(['NCN-KOR-JPN'])
    ws_output.append([])
    ws_output.append(['UPDATE', 'VESSEL', 'DWT', 'BUILT', 'DRAFT', 'DOP', 'LAYDAY', 'OWNER/BROKER'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')

def excelSCHINA(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active

    ws_output.append([])
    ws_output.append(['SCHINA'])
    ws_output.append(['.'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')

def excelSEA(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active

    ws_output.append([])
    ws_output.append(['SEA'])
    ws_output.append(['.'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')

def excelPG(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active


    ws_output.append([])
    ws_output.append(['PG-IND-SAF'])
    ws_output.append(['.'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')

def excelMED(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active

    ws_output.append([])
    ws_output.append(['MED'])
    ws_output.append(['.'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')

def excelECSA(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active


    ws_output.append([])
    ws_output.append(['ECSA'])
    ws_output.append(['.'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')

def excelATL(sortedlist):
    wb_output = openpyxl.load_workbook("TONNAGE LIST.xlsx", data_only = True)
    ws_output = wb_output.active


    ws_output.append([])
    ws_output.append(['ATLANTIC'])
    ws_output.append(['.'])

    for row in sortedlist:
        
        first_eight_columns = row[:8]

        for i, c in enumerate(first_eight_columns):

            column_A = 'A' #0
            column_B = 'B' #1
            column_C = 'C' #3
            column_D = 'D' #4
            column_E = 'E' #5
            column_F = 'F' #6
            column_G = 'G' #7
            column_H = 'H' #7

            last_rowA = len(list(ws_output[column_A]))
            last_rowB = len(list(ws_output[column_B]))
            last_rowC = len(list(ws_output[column_C]))
            last_rowD = len(list(ws_output[column_D]))
            last_rowE = len(list(ws_output[column_E]))
            last_rowF = len(list(ws_output[column_F]))
            last_rowG = len(list(ws_output[column_G]))
            last_rowH = len(list(ws_output[column_H]))



            
            if i == 0 and c.value is not None: #Update 
                formatted_value = module.format_update(c.value)
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 0 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_A + str(last_rowA + 1)] = formatted_value
            elif i == 1 and c.value is not None: #Vessel Name
                formatted_value = str(c.value)
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 1 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_B + str(last_rowB)] = formatted_value
            elif i == 2 and c.value is not None: #DWT
                formatted_value = c.value
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 2 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_C + str(last_rowC)] = formatted_value
            elif i == 3 and c.value is not None: #Built 
                formatted_value = c.value
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 3 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_D + str(last_rowD)] = formatted_value
            elif i == 4 and c.value is not None: #Draft 
                formatted_value = c.value
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 4 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_E + str(last_rowE)] = formatted_value
            elif i == 5 and c.value is not None: #DOP
                formatted_value = str(c.value)
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 5 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_F + str(last_rowF)] = formatted_value
            elif i == 6 and c.value is not None: #LAYDAY
                formatted_value = module.format_layday(c.value)
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 6 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_G + str(last_rowG)] = formatted_value
            elif i == 7 and c.value is not None: #Owner/Broker
                formatted_value = str(c.value)
                ws_output[column_H + str(last_rowH )] = formatted_value
            elif i == 7 and c.value is None:
                formatted_value = "N/A"
                ws_output[column_H + str(last_rowH)] = formatted_value
    wb_output.save('TONNAGE LIST.xlsx')
