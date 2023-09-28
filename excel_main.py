from datetime import date, timedelta
import openpyxl
import excel_module
import module

today = date.today()
dayofweek1 = today.weekday()
dayofweek = int(dayofweek1)

wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
ws_output = wb_output.active


#Monday
if dayofweek == 0:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=3), today - timedelta(days=4)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Tuesday
if dayofweek == 1:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=5), today - timedelta(days=1), today - timedelta(days = 4)] + future_days 

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Wednesday
if dayofweek == 2:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Wednesday
if dayofweek == 3:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Thursday
if dayofweek == 4:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today -timedelta(days=3)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Friday
if dayofweek == 4:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today - timedelta(days=3), today - timedelta(days=4)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Saturday
if dayofweek == 5:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today - timedelta(days=3), today - timedelta(days=4), today - timedelta(days=5)] + future_days


    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)

#Sunday
if dayofweek == 6:

    wb = openpyxl.load_workbook("TONNAGE UPDATE.xlsx", data_only=True)
    ws = wb.active

    wb_output = openpyxl.load_workbook("output1.xlsx", data_only= True)
    ws_output = wb_output.active


    future_days = [today + timedelta(days = i) for i in range(1000)]
    days = [today, today - timedelta(days=2), today - timedelta(days=1), today - timedelta(days=3), today - timedelta(days=4)] + future_days

    rows = module.filter_rows(ws, days)

    for region, rows in rows.items():
        rows.sort(key=lambda row: row[6].value)
        if region == 'NCHINA':
            excel_module.excelNCHINA(rows)
        elif region == 'SCHINA':  
            excel_module.excelSCHINA(rows)
        elif region == 'SEA':
            excel_module.excelSEA(rows)
        elif region == 'PG':
            excel_module.excelPG(rows)
        elif region == 'MED':
            excel_module.excelMED(rows)
        elif region == 'ECSA':
            excel_module.excelECSA(rows)
        elif region == 'BALTIC':
            excel_module.excelATL(rows)